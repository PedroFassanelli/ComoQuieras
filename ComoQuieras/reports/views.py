from django.shortcuts import render
from django.views.generic import TemplateView

from pedidos.models import Pedido, LineaPedido
from Cometela.models import Tamaño

from django.http.response import HttpResponse

from openpyxl import Workbook
from openpyxl.styles import Alignment,Border,Font,PatternFill,Side

from datetime import date

from ComoQuieras.urls import tienda
from django.shortcuts import redirect
from Cometela.views import user_passes_test, es_rrhh


COLUMNAS = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG', 'AH']


class ReportePersonalizadoExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        #Solo usuarios de RRHH pueden secar este reporte
        if es_rrhh(request.user):
            pedidos = Pedido.objects.filter(estado='ACTIVO')
            lineas = LineaPedido.objects.all()
            lineas_pedidos = []
            for pedido in pedidos:
                for linea in lineas:
                    if linea.pedido == pedido:
                        lineas_pedidos.append(linea)
            
            wb = Workbook()
            #Hoja para los pedidos de SUCURSAL
            ws = wb.active
            ws.title = "SUCURSAL"
            #Hoja para los pedidos de DEPOSITO
            ws2 = wb.create_sheet(title="DEPOSITO")

            #Generar los encabezados
            encabezado(ws)
            encabezado(ws2)
            
            #Completar las filas con las cantidades correspondientes a los pedidos
            cont1=3
            cont2=3
            for p in pedidos:  
                if (es_sucursal(p.user)):
                    cont1+=1
                    ws["A" + str(cont1)] = p.user.last_name + ", " + p.user.first_name
                    ws["A" + str(cont1)].alignment = Alignment(horizontal = "center", vertical = "center")   
                    ws["A" + str(cont1)].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                        top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
                    for l in lineas:
                        if l.pedido == p:
                            columna = ""
                            dia = l.vianda_tamaño.vianda.dia
                            tipo = l.vianda_tamaño.vianda.tipo
                            tamaño = l.vianda_tamaño.tamaño.tamaño
                            match dia:
                                case "LUNES":
                                    if tipo == "TRADICIONAL":
                                        if tamaño == "ENTERA":
                                            columna = "B"
                                        else:
                                            columna = "C"
                                    elif tipo == "LIGHT":
                                        if tamaño == "ENTERA":
                                            columna = "D"
                                        else:
                                            columna = "E"
                                    elif tipo == "VEGGIE":
                                        if tamaño == "ENTERA":
                                            columna = "F"
                                        else:
                                            columna = "G"
                                    
                                case "MARTES":
                                    if tipo == "TRADICIONAL":
                                        if tamaño == "ENTERA":
                                            columna = "H"
                                        else:
                                            columna = "I"
                                    elif tipo == "LIGHT":
                                        if tamaño == "ENTERA":
                                            columna = "J"
                                        else:
                                            columna = "K"
                                    elif tipo == "VEGGIE":
                                        if tamaño == "ENTERA":
                                            columna = "L"
                                        else:
                                            columna = "M"
                                    
                                case "MIERCOLES":
                                    if tipo == "TRADICIONAL":
                                        if tamaño == "ENTERA":
                                            columna = "N"
                                        else:
                                            columna = "O"
                                    elif tipo == "LIGHT":
                                        if tamaño == "ENTERA":
                                            columna = "P"
                                        else:
                                            columna = "Q"
                                    elif tipo == "VEGGIE":
                                        if tamaño == "ENTERA":
                                            columna = "R"
                                        else:
                                            columna = "S"
                                
                                case "JUEVES":
                                    if tipo == "TRADICIONAL":
                                        if tamaño == "ENTERA":
                                            columna = "T"
                                        else:
                                            columna = "U"
                                    elif tipo == "LIGHT":
                                        if tamaño == "ENTERA":
                                            columna = "V"
                                        else:
                                            columna = "W"
                                    elif tipo == "VEGGIE":
                                        if tamaño == "ENTERA":
                                            columna = "X"
                                        else:
                                            columna = "Y"
                                
                                case "VIERNES":
                                    if tipo == "TRADICIONAL":
                                        if tamaño == "ENTERA":
                                            columna = "Z"
                                        else:
                                            columna = "AA"
                                    elif tipo == "LIGHT":
                                        if tamaño == "ENTERA":
                                            columna = "AB"
                                        else:
                                            columna = "AC"
                                    elif tipo == "VEGGIE":
                                        if tamaño == "ENTERA":
                                            columna = "AD"
                                        else:
                                            columna = "AE"
                                    
                            ws[columna + str(cont1)] = l.cantidad
                            ws[columna + str(cont1)].alignment = Alignment(horizontal = "center", vertical = "center")
                else:
                    if (es_desposito(p.user)):
                        cont2+=1
                        ws2["A" + str(cont2)] = p.user.last_name + ", " + p.user.first_name
                        ws2["A" + str(cont2)].alignment = Alignment(horizontal = "center", vertical = "center")   
                        ws2["A" + str(cont2)].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
                        for l in lineas:
                            if l.pedido == p:
                                columna = ""
                                dia = l.vianda_tamaño.vianda.dia
                                tipo = l.vianda_tamaño.vianda.tipo
                                tamaño = l.vianda_tamaño.tamaño.tamaño
                                match dia:
                                    case "LUNES":
                                        if tipo == "TRADICIONAL":
                                            if tamaño == "ENTERA":
                                                columna = "B"
                                            else:
                                                columna = "C"
                                        elif tipo == "LIGHT":
                                            if tamaño == "ENTERA":
                                                columna = "D"
                                            else:
                                                columna = "E"
                                        elif tipo == "VEGGIE":
                                            if tamaño == "ENTERA":
                                                columna = "F"
                                            else:
                                                columna = "G"
                                        
                                    case "MARTES":
                                        if tipo == "TRADICIONAL":
                                            if tamaño == "ENTERA":
                                                columna = "H"
                                            else:
                                                columna = "I"
                                        elif tipo == "LIGHT":
                                            if tamaño == "ENTERA":
                                                columna = "J"
                                            else:
                                                columna = "K"
                                        elif tipo == "VEGGIE":
                                            if tamaño == "ENTERA":
                                                columna = "L"
                                            else:
                                                columna = "M"
                                        
                                    case "MIERCOLES":
                                        if tipo == "TRADICIONAL":
                                            if tamaño == "ENTERA":
                                                columna = "N"
                                            else:
                                                columna = "O"
                                        elif tipo == "LIGHT":
                                            if tamaño == "ENTERA":
                                                columna = "P"
                                            else:
                                                columna = "Q"
                                        elif tipo == "VEGGIE":
                                            if tamaño == "ENTERA":
                                                columna = "R"
                                            else:
                                                columna = "S"
                                    
                                    case "JUEVES":
                                        if tipo == "TRADICIONAL":
                                            if tamaño == "ENTERA":
                                                columna = "T"
                                            else:
                                                columna = "U"
                                        elif tipo == "LIGHT":
                                            if tamaño == "ENTERA":
                                                columna = "V"
                                            else:
                                                columna = "W"
                                        elif tipo == "VEGGIE":
                                            if tamaño == "ENTERA":
                                                columna = "X"
                                            else:
                                                columna = "Y"
                                    
                                    case "VIERNES":
                                        if tipo == "TRADICIONAL":
                                            if tamaño == "ENTERA":
                                                columna = "Z"
                                            else:
                                                columna = "AA"
                                        elif tipo == "LIGHT":
                                            if tamaño == "ENTERA":
                                                columna = "AB"
                                            else:
                                                columna = "AC"
                                        elif tipo == "VEGGIE":
                                            if tamaño == "ENTERA":
                                                columna = "AD"
                                            else:
                                                columna = "AE"
                                        
                                ws2[columna + str(cont2)] = l.cantidad
                                ws2[columna + str(cont2)].alignment = Alignment(horizontal = "center", vertical = "center")

            lineas_columnas(ws, cont1)
            lineas_columnas(ws2, cont2)

            #Calcular cantidad de viandas totales y precio por fila(pedido)
            #SUCURSAL
            cont_rows = 0
            precio_media = Tamaño.objects.filter(estado='ACTIVA').filter(tamaño='MEDIA')[0].precio
            precio_entera = Tamaño.objects.filter(estado='ACTIVA').filter(tamaño='ENTERA')[0].precio
            for row in ws.iter_rows(min_row=4, min_col=2, max_col=31, max_row=cont1, values_only=True):
                cont_rows += 1
                contador = 0
                c_medias = 0
                c_enteras = 0
                for x in row:
                    contador += 1
                    if (contador%2 == 0):
                        #Posiciones pares, medias
                        if x is not None:
                            c_medias += x
                    else:
                        #Posiciones impares, enteras
                        if x is not None:
                            c_enteras += x
                ws['AF' + str(cont_rows+3)] = c_enteras
                ws['AF' + str(cont_rows+3)].alignment = Alignment(horizontal = "center", vertical = "center")
                ws['AG' + str(cont_rows+3)] = c_medias
                ws['AG' + str(cont_rows+3)].alignment = Alignment(horizontal = "center", vertical = "center")
                ws['AH' + str(cont_rows+3)] = c_enteras * precio_entera + c_medias * precio_media
                ws['AH' + str(cont_rows+3)].alignment = Alignment(horizontal = "center", vertical = "center")
            ws.column_dimensions['AF'].width = 15
            ws.column_dimensions['AG'].width = 15
            ws.column_dimensions['AH'].width = 10
            #DEPOSITO
            cont_rows2 = 0
            precio_media = Tamaño.objects.filter(estado='ACTIVA').filter(tamaño='MEDIA')[0].precio
            precio_entera = Tamaño.objects.filter(estado='ACTIVA').filter(tamaño='ENTERA')[0].precio
            for row in ws2.iter_rows(min_row=4, min_col=2, max_col=31, max_row=cont2, values_only=True):
                cont_rows2 += 1
                contador2 = 0
                c_medias2 = 0
                c_enteras2 = 0
                for x in row:
                    contador2 += 1
                    if (contador2%2 == 0):
                        #Posiciones pares, medias
                        if x is not None:
                            c_medias2 += x
                    else:
                        #Posiciones impares, enteras
                        if x is not None:
                            c_enteras2 += x
                ws2['AF' + str(cont_rows2+3)] = c_enteras2
                ws2['AF' + str(cont_rows2+3)].alignment = Alignment(horizontal = "center", vertical = "center")
                ws2['AG' + str(cont_rows2+3)] = c_medias2
                ws2['AG' + str(cont_rows2+3)].alignment = Alignment(horizontal = "center", vertical = "center")
                ws2['AH' + str(cont_rows2+3)] = c_enteras2 * precio_entera + c_medias2 * precio_media
                ws2['AH' + str(cont_rows2+3)].alignment = Alignment(horizontal = "center", vertical = "center")
            ws2.column_dimensions['AF'].width = 15
            ws2.column_dimensions['AG'].width = 15
            ws2.column_dimensions['AH'].width = 10

            #Calcular los totales de cada columna
            #SUCURSAL
            cant_cols = 0
            for col in ws.iter_cols(min_row=4, min_col=2, max_col=34, max_row=cont1):
                cant_cols +=1
                cant_filas = 0
                total = 0
                
                for x in col:
                    if x.value is not None:
                        total += x.value
                    cant_filas += 1

                ws[COLUMNAS[cant_cols] + str(cant_filas+4)] = total
            #DEPOSITO
            cant_cols2 = 0
            for col in ws2.iter_cols(min_row=4, min_col=2, max_col=34, max_row=cont2):
                cant_cols2 +=1
                cant_filas2 = 0
                total2 = 0
                
                for x in col:
                    if x.value is not None:
                        total2 += x.value
                    cant_filas2 += 1

                ws2[COLUMNAS[cant_cols2] + str(cant_filas2+4)] = total2

            #Estilo de la fila de totales
            #SUCURSAL
            totales(ws, cant_filas)
            #DEPOSITO
            totales(ws2, cant_filas2)

            nombre_archivo = "PedidoSemanal {0}.xlsx".format(date.today())
            response = HttpResponse(content_type = "application/ms-excel")
            contenido = "attachment; filename = {0}".format(nombre_archivo)
            response["Content-Disposition"] = contenido
            wb.save(response)
            return response
        else:
            return redirect('Tienda')


#Controlar a que grupo pertenece el usuario
def es_sucursal(user):
    return user.groups.filter(name__in=['SUCURSAL'])

def es_desposito(user):
    return user.groups.filter(name__in=['DEPOSITO'])

#Formato tabla
def encabezado(ws):
    #Usuario
    ws['A1'].alignment = Alignment(horizontal = "center", vertical = "center")    
    ws['A1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style= "thin"),
                            top= Side(border_style= "thin"), bottom= Side(border_style= "thin"))
    ws['A1'].fill = PatternFill(start_color= '66CFCC', end_color='66CFCC', fill_type= "solid")
    ws['A1'].font = Font(name = 'Colibri', size=12, bold=True)
    ws['A1'] = 'Usuario'
    ws.merge_cells('A1:A3')
    
    ws.column_dimensions['A'].width = 30
    #Fijamos las primer columna y las 3 primeras filas
    ws.freeze_panes = ws['B4']

    #Dias
    ws['B1'].alignment = Alignment(horizontal = "center",vertical = "center")
    ws['B1'].border = Border(left = Side(border_style = "medium"), right = Side(border_style = "medium"),
                            top = Side(border_style = "medium"), bottom = Side(border_style = "thin") ) 
    ws['B1'].fill = PatternFill(start_color = 'FFF2CC', end_color = 'FFF2CC', fill_type = "solid")
    ws['B1'].font = Font(name = 'Calibri', size = 12, bold = True)
    ws['B1'] = 'LUNES'
    ws.merge_cells('B1:G1')

    ws['H1'].alignment = Alignment(horizontal = "center",vertical = "center")
    ws['H1'].border = Border(left = Side(border_style = "medium"), right = Side(border_style = "medium"),
                            top = Side(border_style = "medium"), bottom = Side(border_style = "thin") ) 
    ws['H1'].fill = PatternFill(start_color = 'C6E0B4', end_color = 'C6E0B4', fill_type = "solid")
    ws['H1'].font = Font(name = 'Calibri', size = 12, bold = True)
    ws['H1'] = 'MARTES'
    ws.merge_cells('H1:M1')

    ws['N1'].alignment = Alignment(horizontal = "center",vertical = "center")
    ws['N1'].border = Border(left = Side(border_style = "medium"), right = Side(border_style = "medium"),
                            top = Side(border_style = "medium"), bottom = Side(border_style = "thin") ) 
    ws['N1'].fill = PatternFill(start_color = 'F8CBAD', end_color = 'F8CBAD', fill_type = "solid")
    ws['N1'].font = Font(name = 'Calibri', size = 12, bold = True)
    ws['N1'] = 'MIERCOLES'
    ws.merge_cells('N1:S1')

    ws['T1'].alignment = Alignment(horizontal = "center",vertical = "center")
    ws['T1'].border = Border(left = Side(border_style = "medium"), right = Side(border_style = "medium"),
                            top = Side(border_style = "medium"), bottom = Side(border_style = "thin") ) 
    ws['T1'].fill = PatternFill(start_color = 'D9E2F3', end_color = 'D9E2F3', fill_type = "solid")
    ws['T1'].font = Font(name = 'Calibri', size = 12, bold = True)
    ws['T1'] = 'JUEVES'
    ws.merge_cells('T1:Y1')

    ws['Z1'].alignment = Alignment(horizontal = "center",vertical = "center")
    ws['Z1'].border = Border(left = Side(border_style = "medium"), right = Side(border_style = "medium"),
                            top = Side(border_style = "medium"), bottom = Side(border_style = "thin") ) 
    ws['Z1'].fill = PatternFill(start_color = 'FFB9B9', end_color = 'FFB9B9', fill_type = "solid")
    ws['Z1'].font = Font(name = 'Calibri', size = 12, bold = True)
    ws['Z1'] = 'VIERNES'
    ws.merge_cells('Z1:AE1')

    ws['AF1'].alignment = Alignment(horizontal = "center",vertical = "center")
    ws['AF1'].border = Border(left = Side(border_style = "medium"), right = Side(border_style = "medium"),
                            top = Side(border_style = "medium"), bottom = Side(border_style = "medium") ) 
    ws['AF1'].fill = PatternFill(start_color = 'DACACA', end_color = 'DACACA', fill_type = "solid")
    ws['AF1'].font = Font(name = 'Calibri', size = 12, bold = True)
    ws['AF1'] = 'TOTAL SEMANAL PERSONA'
    ws.merge_cells('AF1:AH2')

    #Tipos
    ws['B2'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['B2'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "medium") )
    ws['B2'].fill = PatternFill(start_color = 'FFF2CC', end_color = 'FFF2CC', fill_type = "solid")
    ws['B2'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['B2'] = 'TRADICIONAL'
    ws.merge_cells('B2:C2')
    ws['D2'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['D2'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "medium") )
    ws['D2'].fill = PatternFill(start_color = 'FFF2CC', end_color = 'FFF2CC', fill_type = "solid")
    ws['D2'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['D2'] = 'LIGHT'
    ws.merge_cells('D2:E2')
    ws['F2'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['F2'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "medium"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "medium") )
    ws['F2'].fill = PatternFill(start_color = 'FFF2CC', end_color = 'FFF2CC', fill_type = "solid")
    ws['F2'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['F2'] = 'VEGGIE'
    ws.merge_cells('F2:G2')
    ws['H2'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['H2'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "medium") )
    ws['H2'].fill = PatternFill(start_color = 'C6E0B4', end_color = 'C6E0B4', fill_type = "solid")
    ws['H2'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['H2'] = 'TRADICIONAL'
    ws.merge_cells('H2:I2')
    ws['J2'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['J2'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "medium") )
    ws['J2'].fill = PatternFill(start_color = 'C6E0B4', end_color = 'C6E0B4', fill_type = "solid")
    ws['J2'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['J2'] = 'LIGHT'
    ws.merge_cells('J2:K2')
    ws['L2'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['L2'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "medium"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "medium") )
    ws['L2'].fill = PatternFill(start_color = 'C6E0B4', end_color = 'C6E0B4', fill_type = "solid")
    ws['L2'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['L2'] = 'VEGGIE'
    ws.merge_cells('L2:M2')
    ws['N2'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['N2'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "medium") )
    ws['N2'].fill = PatternFill(start_color = 'F8CBAD', end_color = 'F8CBAD', fill_type = "solid")
    ws['N2'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['N2'] = 'TRADICIONAL'
    ws.merge_cells('N2:O2')
    ws['P2'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['P2'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "medium") )
    ws['P2'].fill = PatternFill(start_color = 'F8CBAD', end_color = 'F8CBAD', fill_type = "solid")
    ws['P2'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['P2'] = 'LIGHT'
    ws.merge_cells('P2:Q2')
    ws['R2'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['R2'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "medium"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "medium") )
    ws['R2'].fill = PatternFill(start_color = 'F8CBAD', end_color = 'F8CBAD', fill_type = "solid")
    ws['R2'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['R2'] = 'VEGGIE'
    ws.merge_cells('R2:S2')
    ws['T2'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['T2'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "medium") )
    ws['T2'].fill = PatternFill(start_color = 'D9E2F3', end_color = 'D9E2F3', fill_type = "solid")
    ws['T2'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['T2'] = 'TRADICIONAL'
    ws.merge_cells('T2:U2')
    ws['V2'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['V2'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "medium") )
    ws['V2'].fill = PatternFill(start_color = 'D9E2F3', end_color = 'D9E2F3', fill_type = "solid")
    ws['V2'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['V2'] = 'LIGHT'
    ws.merge_cells('V2:W2')
    ws['X2'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['X2'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "medium"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "medium") )
    ws['X2'].fill = PatternFill(start_color = 'D9E2F3', end_color = 'D9E2F3', fill_type = "solid")
    ws['X2'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['X2'] = 'VEGGIE'
    ws.merge_cells('X2:Y2')
    ws['Z2'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['Z2'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "medium") )
    ws['Z2'].fill = PatternFill(start_color = 'FFB9B9', end_color = 'FFB9B9', fill_type = "solid")
    ws['Z2'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['Z2'] = 'TRADICIONAL'
    ws.merge_cells('Z2:AA2')
    ws['AB2'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['AB2'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "medium") )
    ws['AB2'].fill = PatternFill(start_color = 'FFB9B9', end_color = 'FFB9B9', fill_type = "solid")
    ws['AB2'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['AB2'] = 'LIGHT'
    ws.merge_cells('AB2:AC2')
    ws['AD2'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['AD2'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "medium"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "medium") )
    ws['AD2'].fill = PatternFill(start_color = 'FFB9B9', end_color = 'FFB9B9', fill_type = "solid")
    ws['AD2'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['AD2'] = 'VEGGIE'
    ws.merge_cells('AD2:AE2')

    #Tamaños
    ws['B3'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['B3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
    ws['B3'].fill = PatternFill(start_color = 'FFF2CC', end_color = 'FFF2CC', fill_type = "solid")
    ws['B3'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['B3'] = 'ENTERA'
    ws['C3'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['C3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
    ws['C3'].fill = PatternFill(start_color = 'FFF2CC', end_color = 'FFF2CC', fill_type = "solid")
    ws['C3'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['C3'] = 'MEDIA'
    ws['D3'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['D3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
    ws['D3'].fill = PatternFill(start_color = 'FFF2CC', end_color = 'FFF2CC', fill_type = "solid")
    ws['D3'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['D3'] = 'ENTERA'
    ws['E3'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['E3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
    ws['E3'].fill = PatternFill(start_color = 'FFF2CC', end_color = 'FFF2CC', fill_type = "solid")
    ws['E3'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['E3'] = 'MEDIA'
    ws['F3'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['F3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
    ws['F3'].fill = PatternFill(start_color = 'FFF2CC', end_color = 'FFF2CC', fill_type = "solid")
    ws['F3'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['F3'] = 'ENTERA'
    ws['G3'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['G3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "medium"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
    ws['G3'].fill = PatternFill(start_color = 'FFF2CC', end_color = 'FFF2CC', fill_type = "solid")
    ws['G3'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['G3'] = 'MEDIA'
    ws['H3'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['H3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
    ws['H3'].fill = PatternFill(start_color = 'C6E0B4', end_color = 'C6E0B4', fill_type = "solid")
    ws['H3'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['H3'] = 'ENTERA'
    ws['I3'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['I3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
    ws['I3'].fill = PatternFill(start_color = 'C6E0B4', end_color = 'C6E0B4', fill_type = "solid")
    ws['I3'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['I3'] = 'MEDIA'
    ws['J3'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['J3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
    ws['J3'].fill = PatternFill(start_color = 'C6E0B4', end_color = 'C6E0B4', fill_type = "solid")
    ws['J3'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['J3'] = 'ENTERA'
    ws['K3'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['K3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
    ws['K3'].fill = PatternFill(start_color = 'C6E0B4', end_color = 'C6E0B4', fill_type = "solid")
    ws['K3'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['K3'] = 'MEDIA'
    ws['L3'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['L3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
    ws['L3'].fill = PatternFill(start_color = 'C6E0B4', end_color = 'C6E0B4', fill_type = "solid")
    ws['L3'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['L3'] = 'ENTERA'
    ws['M3'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['M3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "medium"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
    ws['M3'].fill = PatternFill(start_color = 'C6E0B4', end_color = 'C6E0B4', fill_type = "solid")
    ws['M3'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['M3'] = 'MEDIA'
    ws['N3'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['N3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
    ws['N3'].fill = PatternFill(start_color = 'F8CBAD', end_color = 'F8CBAD', fill_type = "solid")
    ws['N3'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['N3'] = 'ENTERA'
    ws['O3'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['O3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
    ws['O3'].fill = PatternFill(start_color = 'F8CBAD', end_color = 'F8CBAD', fill_type = "solid")
    ws['O3'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['O3'] = 'MEDIA'
    ws['P3'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['P3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
    ws['P3'].fill = PatternFill(start_color = 'F8CBAD', end_color = 'F8CBAD', fill_type = "solid")
    ws['P3'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['P3'] = 'ENTERA'
    ws['Q3'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['Q3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
    ws['Q3'].fill = PatternFill(start_color = 'F8CBAD', end_color = 'F8CBAD', fill_type = "solid")
    ws['Q3'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['Q3'] = 'MEDIA'
    ws['R3'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['R3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
    ws['R3'].fill = PatternFill(start_color = 'F8CBAD', end_color = 'F8CBAD', fill_type = "solid")
    ws['R3'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['R3'] = 'ENTERA'
    ws['S3'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['S3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "medium"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
    ws['S3'].fill = PatternFill(start_color = 'F8CBAD', end_color = 'F8CBAD', fill_type = "solid")
    ws['S3'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['S3'] = 'MEDIA'
    ws['T3'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['T3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
    ws['T3'].fill = PatternFill(start_color = 'D9E2F3', end_color = 'D9E2F3', fill_type = "solid")
    ws['T3'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['T3'] = 'ENTERA'
    ws['U3'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['U3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
    ws['U3'].fill = PatternFill(start_color = 'D9E2F3', end_color = 'D9E2F3', fill_type = "solid")
    ws['U3'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['U3'] = 'MEDIA'
    ws['V3'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['V3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
    ws['V3'].fill = PatternFill(start_color = 'D9E2F3', end_color = 'D9E2F3', fill_type = "solid")
    ws['V3'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['V3'] = 'ENTERA'
    ws['W3'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['W3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
    ws['W3'].fill = PatternFill(start_color = 'D9E2F3', end_color = 'D9E2F3', fill_type = "solid")
    ws['W3'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['W3'] = 'MEDIA'
    ws['X3'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['X3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
    ws['X3'].fill = PatternFill(start_color = 'D9E2F3', end_color = 'D9E2F3', fill_type = "solid")
    ws['X3'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['X3'] = 'ENTERA'
    ws['Y3'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['Y3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "medium"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
    ws['Y3'].fill = PatternFill(start_color = 'D9E2F3', end_color = 'D9E2F3', fill_type = "solid")
    ws['Y3'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['Y3'] = 'MEDIA'
    ws['Z3'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['Z3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
    ws['Z3'].fill = PatternFill(start_color = 'FFB9B9', end_color = 'FFB9B9', fill_type = "solid")
    ws['Z3'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['Z3'] = 'ENTERA'
    ws['AA3'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['AA3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
    ws['AA3'].fill = PatternFill(start_color = 'FFB9B9', end_color = 'FFB9B9', fill_type = "solid")
    ws['AA3'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['AA3'] = 'MEDIA'
    ws['AB3'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['AB3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
    ws['AB3'].fill = PatternFill(start_color = 'FFB9B9', end_color = 'FFB9B9', fill_type = "solid")
    ws['AB3'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['AB3'] = 'ENTERA'
    ws['AC3'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['AC3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
    ws['AC3'].fill = PatternFill(start_color = 'FFB9B9', end_color = 'FFB9B9', fill_type = "solid")
    ws['AC3'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['AC3'] = 'MEDIA'
    ws['AD3'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['AD3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
    ws['AD3'].fill = PatternFill(start_color = 'FFB9B9', end_color = 'FFB9B9', fill_type = "solid")
    ws['AD3'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['AD3'] = 'ENTERA'
    ws['AE3'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['AE3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "medium"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
    ws['AE3'].fill = PatternFill(start_color = 'FFB9B9', end_color = 'FFB9B9', fill_type = "solid")
    ws['AE3'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['AE3'] = 'MEDIA'
    ws['AF3'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['AF3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
    ws['AF3'].fill = PatternFill(start_color = 'DACACA', end_color = 'DACACA', fill_type = "solid")
    ws['AF3'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['AF3'] = 'Suma Enteras'
    ws['AG3'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['AG3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
    ws['AG3'].fill = PatternFill(start_color = 'DACACA', end_color = 'DACACA', fill_type = "solid")
    ws['AG3'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['AG3'] = 'Suma Medias'
    ws['AH3'].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['AH3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "medium"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
    ws['AH3'].fill = PatternFill(start_color = 'DACACA', end_color = 'DACACA', fill_type = "solid")
    ws['AH3'].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['AH3'] = 'Valor'


def totales(ws, cant_filas):  
    ws['A' + str(cant_filas+4)].alignment = Alignment(horizontal = "center", vertical = "center")
    ws['A'+ str(cant_filas+4)].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
    ws['A'+ str(cant_filas+4)].fill = PatternFill(start_color = 'DACACA', end_color = 'DACACA', fill_type = "solid")
    ws['A'+ str(cant_filas+4)].font = Font(name = 'Calibro', size = 10, bold = True)
    ws['A'+ str(cant_filas+4)] = 'TOTALES'

    for c in COLUMNAS:
        if c != 'A':
            ws[c + str(cant_filas+4)].alignment = Alignment(horizontal = "center", vertical = "center")
            ws[c + str(cant_filas+4)].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                            top = Side(border_style = "medium"), bottom = Side(border_style = "medium") )
            ws[c + str(cant_filas+4)].font = Font(name = 'Calibro', size = 10, bold = True)
            if c == 'B' or c == 'C' or c == 'D' or c == 'E' or c == 'F' or c == 'G':
                ws[c + str(cant_filas+4)].fill = PatternFill(start_color = 'FFF2CC', end_color = 'FFF2CC', fill_type = "solid")
            elif c == 'H' or c == 'I' or c == 'J' or c == 'K' or c == 'L' or c == 'M':
                ws[c + str(cant_filas+4)].fill = PatternFill(start_color = 'C6E0B4', end_color = 'C6E0B4', fill_type = "solid")
            elif c == 'N' or c == 'O' or c == 'P' or c == 'Q' or c == 'R' or c == 'S':
                ws[c + str(cant_filas+4)].fill = PatternFill(start_color = 'F8CBAD', end_color = 'F8CBAD', fill_type = "solid")
            elif c == 'T' or c == 'U' or c == 'V' or c == 'W' or c == 'X' or c == 'Y':
                ws[c + str(cant_filas+4)].fill = PatternFill(start_color = 'D9E2F3', end_color = 'D9E2F3', fill_type = "solid")
            elif c == 'Z' or c == 'AA' or c == 'AB' or c == 'AC' or c == 'AD' or c == 'AE':   
                ws[c + str(cant_filas+4)].fill = PatternFill(start_color = 'FFB9B9', end_color = 'FFB9B9', fill_type = "solid")
            else:
                ws[c + str(cant_filas+4)].fill = PatternFill(start_color = 'DACACA', end_color = 'DACACA', fill_type = "solid")
    
    ws['G' + str(cant_filas+4)].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "medium"),
                            top = Side(border_style = "medium"), bottom = Side(border_style = "medium") )
    ws['M' + str(cant_filas+4)].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "medium"),
                            top = Side(border_style = "medium"), bottom = Side(border_style = "medium") )    
    ws['S' + str(cant_filas+4)].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "medium"),
                            top = Side(border_style = "medium"), bottom = Side(border_style = "medium") )
    ws['Y' + str(cant_filas+4)].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "medium"),
                            top = Side(border_style = "medium"), bottom = Side(border_style = "medium") )
    ws['AE' + str(cant_filas+4)].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "medium"),
                            top = Side(border_style = "medium"), bottom = Side(border_style = "medium") )
    ws['AH' + str(cant_filas+4)].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "medium"),
                            top = Side(border_style = "medium"), bottom = Side(border_style = "medium") )



def lineas_columnas(ws, cont):
    for c in range(cont-3):
        for col in COLUMNAS:
            ws[col + str(c+4)].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"))
        ws['G' + str(c+4)].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "medium"))
        ws['M' + str(c+4)].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "medium"))
        ws['S' + str(c+4)].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "medium"))
        ws['Y' + str(c+4)].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "medium"))
        ws['AE' + str(c+4)].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "medium"))
        ws['AH' + str(c+4)].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "medium"))